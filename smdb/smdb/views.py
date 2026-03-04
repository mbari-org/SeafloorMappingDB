import json
import logging
import re
import csv
from os.path import join
from io import BytesIO
from django.utils.dateparse import parse_date

from django.conf import settings
from django.contrib import messages
from django.contrib.messages.views import SuccessMessageMixin
from django.contrib.gis.geos import Polygon
from django.db import connection
from django.db.models import Max, Min, Q
from django.http import HttpResponse, JsonResponse
from django.views import View
from django.views.generic import DetailView
from django.views.generic.base import TemplateView
from django_filters.views import FilterView
from django_tables2 import SingleTableView, RequestConfig
from rest_framework_gis.serializers import (
    GeoFeatureModelSerializer,
    GeometrySerializerMethodField,
)
from rest_framework.serializers import HyperlinkedModelSerializer

from smdb.filters import CompilationFilter, ExpeditionFilter, MissionFilter
from smdb.forms import (
    CompilationFilterFormHelper,
    CompilationFilterSidebarHelper,
    ExpeditionFilterFormHelper,
    ExpeditionFilterSidebarHelper,
    MissionFilterFormHelper,
    MissionFilterSidebarHelper,
)
from smdb.models import Compilation, Expedition, Mission, MBARI_DIR
from smdb.tables import CompilationTable, ExpeditionTable, MissionTable


class ExpeditionSerializer(HyperlinkedModelSerializer):
    class Meta:
        model = Expedition
        fields = ("name",)


class MissionSerializer(GeoFeatureModelSerializer):
    """Should probably be in smdb.api.base.serializers with a complete
    list fields, but here we have it just meeting our needs for MissionOverView()."""

    expedition = ExpeditionSerializer()

    class Meta:
        model = Mission
        geo_field = "nav_track"  # Use nav_track to display track lines, not bounding boxes
        fields = (
            "slug",
            "name",
            "thumbnail_image",
            "start_date",
            "start_ems",
            "end_ems",
            "expedition",
            "route_file",
        )
        nav_track = GeometrySerializerMethodField()
    
    def to_representation(self, instance):
        """Only return missions with nav_track (track lines). Skip missions with only grid_bounds."""
        # Only display missions that have nav_track data (track lines)
        # Check if nav_track exists and is not empty
        if instance.nav_track is None or instance.nav_track.empty:
            # Skip missions without nav_track - do not display bounding boxes
            return None
        
        # Use nav_track for track lines
        return super().to_representation(instance)


def _parse_bbox_geom(request):
    """Parse xmin/xmax/ymin/ymax from a request's GET params into a Polygon.

    Returns None if any parameter is absent, non-numeric, out of valid
    geographic range (-180≤lon≤180, -90≤lat≤90), or inverted (min > max).

    Uses strict ±180 longitude (WGS84) and rejects out-of-range values instead
    of clamping. MissionSelectAPIView/MissionExportAPIView accept up to ±360
    then clamp to ±180; this path is used by MissionTableView (sidebar bbox)
    and intentionally uses strict bounds so invalid/wrapped coords yield no
    results rather than potentially confusing clamped results.
    """
    # Require all four bbox parameters to be present before parsing.
    if not all(request.GET.get(k) for k in ("xmin", "xmax", "ymin", "ymax")):
        return None
    try:
        min_lon = float(request.GET["xmin"])
        max_lon = float(request.GET["xmax"])
        min_lat = float(request.GET["ymin"])
        max_lat = float(request.GET["ymax"])
    except (TypeError, ValueError):
        return None
    # Strict WGS84 bounds: reject lon/lat outside ±180/±90 (see docstring).
    if not (
        -180.0 <= min_lon <= 180.0
        and -180.0 <= max_lon <= 180.0
        and -90.0 <= min_lat <= 90.0
        and -90.0 <= max_lat <= 90.0
        and min_lon <= max_lon
        and min_lat <= max_lat
    ):
        return None
    return Polygon(
        (
            (min_lon, min_lat),
            (min_lon, max_lat),
            (max_lon, max_lat),
            (max_lon, min_lat),
            (min_lon, min_lat),
        ),
        srid=4326,
    )


class MissionOverView(TemplateView):
    logger = logging.getLogger(__name__)
    template_name = "pages/home.html"

    def get_context_data(self, **kwargs):
        """Return the view context data."""
        context = super().get_context_data(**kwargs)
        
        # Add all filter forms to context for sidebar (Missions, Expeditions, Compilations)
        # Use sidebar-specific helpers for vertical layout in the sidebar
        missions_queryset = Mission.objects.all()
        mission_filter = MissionFilter(self.request.GET, queryset=missions_queryset)
        mission_filter.form.helper = MissionFilterSidebarHelper()
        context["mission_filter"] = mission_filter
        
        expeditions_queryset = Expedition.objects.all()
        expedition_filter = ExpeditionFilter(self.request.GET, queryset=expeditions_queryset)
        expedition_filter.form.helper = ExpeditionFilterSidebarHelper()
        context["expedition_filter"] = expedition_filter
        
        compilations_queryset = Compilation.objects.all()
        compilation_filter = CompilationFilter(self.request.GET, queryset=compilations_queryset)
        compilation_filter.form.helper = CompilationFilterSidebarHelper()
        context["compilation_filter"] = compilation_filter
        
        # Default to mission filter for backward compatibility
        context["filter"] = mission_filter
        
        search_string = context["view"].request.GET.get("q")
        search_geom = _parse_bbox_geom(context["view"].request)
        # Check which filter type is active based on filter_type parameter or field presence
        filter_type = self.request.GET.get('filter_type', '')
        
        # Check if any filter parameters are present in the request
        has_mission_filter_params = any(
            key in self.request.GET and self.request.GET.get(key)
            for key in [
                'region_name', 'vehicle_name', 'platformtype',
                'quality_categories', 'patch_test', 'repeat_survey',
                'mgds_compilation', 'expedition__name', 'citation', 'citation_search',
            ]
        ) or (filter_type == 'mission' and 'name' in self.request.GET and self.request.GET.get('name'))
        
        has_expedition_filter_params = filter_type == 'expedition' or (
            'name' in self.request.GET and self.request.GET.get('name') 
            and not has_mission_filter_params and filter_type != 'compilation'
        )
        
        has_compilation_filter_params = filter_type == 'compilation' or (
            'name' in self.request.GET and self.request.GET.get('name') 
            and not has_mission_filter_params and filter_type != 'expedition'
        )
        
        # Start with base queryset
        missions = Mission.objects.order_by("start_date").all()
        
        # Apply mission filter if mission filter parameters are present
        if has_mission_filter_params:
            missions = mission_filter.qs.order_by("start_date")
        
        # Apply expedition filter if expedition filter parameters are present
        elif has_expedition_filter_params:
            filtered_expeditions = expedition_filter.qs
            if filtered_expeditions.exists():
                missions = missions.filter(expedition__in=filtered_expeditions).order_by("start_date")
        
        # Apply compilation filter if compilation filter parameters are present
        elif has_compilation_filter_params:
            filtered_compilations = compilation_filter.qs
            if filtered_compilations.exists():
                missions = missions.filter(compilations__in=filtered_compilations).distinct().order_by("start_date")
        
        # Apply additional filters (search string, geometry, time)
        if search_string:
            self.logger.info("search_string = %s", search_string)
            missions = missions.filter(
                Q(name__icontains=search_string)
                | Q(notes_text__icontains=search_string)
                | Q(expedition__name__icontains=search_string)
            )
        if search_geom:
            # Same spatial logic as MissionTableView/API: nav_track or start_point in bbox
            # (exclude grid_bounds — it can be larger than the track and would return missions
            # whose tracks don't actually pass through the drawn box).
            missions = missions.filter(
                Q(nav_track__intersects=search_geom)
                | Q(start_point__within=search_geom)
            )
        if context["view"].request.GET.get("tmin"):
            min_date_str = context["view"].request.GET.get("tmin")
            max_date_str = context["view"].request.GET.get("tmax")
            # Parse string dates to date objects to avoid TypeError with datetime.combine()
            min_date = parse_date(str(min_date_str)) if min_date_str else None
            max_date = parse_date(str(max_date_str)) if max_date_str else None
            if min_date and max_date:
                missions = missions.filter(
                    start_date__gte=min_date,
                    end_date__lte=max_date,
                )

        # Ensure nav_track and expedition are selected for serialization
        # Filter to only missions with nav_track at the database level for efficiency
        missions = missions.filter(nav_track__isnull=False).exclude(nav_track__isempty=True).select_related("expedition")

        self.logger.info(
            "Serializing %s missions to geojson...",
            missions.count(),
        )
        context["missions"] = MissionSerializer(missions, many=True).data
        self.logger.debug("# of Queries: %d", len(connection.queries))
        self.logger.debug(
            "Size of context['missions']: %d", len(str(context["missions"]))
        )
        self.logger.debug(
            "context['missions'] = %s",
            json.dumps(context["missions"], indent=4, sort_keys=True),
        )
        context["num_missions"] = len(context["missions"]["features"])
        if search_string:
            context["search_string"] = f"containing '{search_string}'"
        time_bounds = missions.aggregate(Min("start_date"), Max("end_date"))
        if time_bounds["start_date__min"]:
            context["start_ems"] = time_bounds["start_date__min"].timestamp() * 1000.0
        if time_bounds["end_date__max"]:
            context["end_ems"] = time_bounds["end_date__max"].timestamp() * 1000.0
        return context


class CompilationTableView(FilterView, SingleTableView):
    table_class = CompilationTable
    queryset = Compilation.objects.all().order_by("name")
    filterset_class = CompilationFilter
    formhelper_class = CompilationFilterFormHelper

    def get_filterset(self, filterset_class):
        kwargs = self.get_filterset_kwargs(filterset_class)
        filterset = filterset_class(**kwargs)
        filterset.form.helper = self.formhelper_class()
        return filterset

    def get_context_data(self, *args, **kwargs):
        # Call the base implementation first to get a context - then add filtered Missions
        context = super().get_context_data(**kwargs)
        
        # Add all filter forms to context for sidebar (Missions, Expeditions, Compilations)
        missions_queryset = Mission.objects.all()
        mission_filter = MissionFilter(self.request.GET, queryset=missions_queryset)
        mission_filter.form.helper = MissionFilterFormHelper()
        context["mission_filter"] = mission_filter
        
        expeditions_queryset = Expedition.objects.all()
        expedition_filter = ExpeditionFilter(self.request.GET, queryset=expeditions_queryset)
        expedition_filter.form.helper = ExpeditionFilterFormHelper()
        context["expedition_filter"] = expedition_filter
        
        compilations_queryset = Compilation.objects.all()
        compilation_filter = CompilationFilter(self.request.GET, queryset=compilations_queryset)
        compilation_filter.form.helper = CompilationFilterFormHelper()
        context["compilation_filter"] = compilation_filter
        
        compilations = CompilationFilter(
            self.request.GET, queryset=self.get_queryset()
        ).qs
        sort = self.request.GET.get("sort")
        if sort:
            compilations = compilations.order_by(sort)
        per_page = int(self.request.GET.get("per_page", 10))
        page = int(self.request.GET.get("page", 1))
        compilations = compilations[slice((page - 1) * per_page, page * per_page)]
        missions = Mission.objects.all()
        missions = (
            missions.filter(compilations__in=compilations)
            .select_related("expedition")
            .distinct()
        )
        # Filter to only missions with nav_track (for map display)
        # This ensures the map shows track lines, not just bounding boxes
        missions = missions.filter(nav_track__isnull=False).exclude(nav_track__isempty=True)
        context["missions"] = MissionSerializer(missions, many=True).data
        return context


class ExpeditionTableView(FilterView, SingleTableView):
    table_class = ExpeditionTable
    queryset = Expedition.objects.all().order_by("name")
    filterset_class = ExpeditionFilter
    filterhelper_class = ExpeditionFilterFormHelper

    def get_filterset(self, filterset_class):
        kwargs = self.get_filterset_kwargs(filterset_class)
        filterset = filterset_class(**kwargs)
        filterset.form.helper = ExpeditionFilterFormHelper()
        return filterset

    def get_context_data(self, *args, **kwargs):
        # Call the base implementation first to get a context - then add filtered Missions
        context = super().get_context_data(**kwargs)
        
        # Add all filter forms to context for sidebar (Missions, Expeditions, Compilations)
        missions_queryset = Mission.objects.all()
        mission_filter = MissionFilter(self.request.GET, queryset=missions_queryset)
        mission_filter.form.helper = MissionFilterFormHelper()
        context["mission_filter"] = mission_filter
        
        expeditions_queryset = Expedition.objects.all()
        expedition_filter = ExpeditionFilter(self.request.GET, queryset=expeditions_queryset)
        expedition_filter.form.helper = ExpeditionFilterFormHelper()
        context["expedition_filter"] = expedition_filter
        
        compilations_queryset = Compilation.objects.all()
        compilation_filter = CompilationFilter(self.request.GET, queryset=compilations_queryset)
        compilation_filter.form.helper = CompilationFilterFormHelper()
        context["compilation_filter"] = compilation_filter
        
        expeditions = ExpeditionFilter(
            self.request.GET, queryset=self.get_queryset()
        ).qs
        sort = self.request.GET.get("sort")
        if sort:
            expeditions = expeditions.order_by(sort)
        per_page = int(self.request.GET.get("per_page", 10))
        page = int(self.request.GET.get("page", 1))
        expeditions = expeditions[slice((page - 1) * per_page, page * per_page)]
        missions = Mission.objects.all()
        missions = (
            missions.filter(expedition__in=expeditions)
            .select_related("expedition")
            .distinct()
        )
        # Filter to only missions with nav_track (for map display)
        # This ensures the map shows track lines, not just bounding boxes
        missions = missions.filter(nav_track__isnull=False).exclude(nav_track__isempty=True)
        context["missions"] = MissionSerializer(missions, many=True).data
        return context


class MissionTableView(FilterView, SingleTableView):
    table_class = MissionTable
    queryset = Mission.objects.all().order_by("name")
    filterset_class = MissionFilter
    formhelper_class = MissionFilterSidebarHelper  # sidebar layout for the collapsible panel

    def _get_bbox_geom(self):
        """Delegate to the module-level _parse_bbox_geom helper."""
        return _parse_bbox_geom(self.request)

    def get_queryset(self):
        """Return base queryset with optional bbox pre-filter applied.

        Applying the bbox here ensures both the rendered table (via FilterView)
        and the map serializer in get_context_data() operate on the same rows.

        If all four bbox params are present but the bbox is invalid (e.g. inverted
        coords), return an empty queryset. Partial bbox (e.g. only xmin) is
        ignored so the table is not emptied by mistake.
        """
        qs = Mission.objects.select_related("expedition").all().order_by("name")
        if all(self.request.GET.get(k) for k in ("xmin", "xmax", "ymin", "ymax")):
            search_geom = self._get_bbox_geom()
            if search_geom:
                qs = qs.filter(
                    Q(nav_track__intersects=search_geom)
                    | Q(start_point__within=search_geom)
                )
            else:
                # xmin was supplied but coords are invalid — return nothing.
                qs = qs.none()
        return qs

    def get_filterset(self, filterset_class):
        kwargs = self.get_filterset_kwargs(filterset_class)
        filterset = filterset_class(**kwargs)
        filterset.form.helper = self.formhelper_class()
        return filterset

    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(**kwargs)

        # self.object_list is the filterset-filtered queryset set by FilterView
        # before get_context_data() is called; it already incorporates both the
        # MissionFilter fields and the bbox pre-filter from get_queryset().
        missions = self.object_list

        sort = self.request.GET.get("sort")
        if sort:
            missions = missions.order_by(sort)

        # Only pass missions with track lines to the map serializer.
        missions = missions.filter(
            nav_track__isnull=False
        ).exclude(nav_track__isempty=True)

        # Map GeoJSON pagination: default 500, cap 2000 (issue #293). Use map_per_page so it does
        # not conflict with django_tables2's per_page (table pagination, e.g. 25 per page). The map
        # intentionally shows a larger slice so many tracks are visible; table pages beyond this
        # slice may show rows that have no track on the map.
        try:
            map_per_page = int(self.request.GET.get("map_per_page", 500))
        except (TypeError, ValueError):
            map_per_page = 500
        map_per_page = max(1, min(map_per_page, 2000))

        # Use a dedicated query parameter for map pagination to avoid conflicts with
        # django_tables2's `page` parameter used for table pagination.
        try:
            map_page = int(self.request.GET.get("map_page", 1))
        except (TypeError, ValueError):
            map_page = 1
        map_page = max(1, map_page)

        missions = missions[slice((map_page - 1) * map_per_page, map_page * map_per_page)]
        context["missions"] = MissionSerializer(missions, many=True).data
        return context


class MissionDetailView(DetailView):
    model = Mission
    queryset = Mission.objects.prefetch_related("citations").all()

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        mission = self.get_object()
        try:
            context["thumbnail_url"] = mission.thumbnail_image.url
        except (AttributeError, ValueError):
            context["thumbnail_url"] = join(
                settings.STATIC_URL, "images", "No_ZTopoSlopeNav_image.jpg"
            )
        try:
            site_uri = self.request.build_absolute_uri("/")
            base_uri = re.match(r"(http[s]*:\/\/[^:\/]*)", site_uri).group(1)
            img_path = mission.thumbnail_filename.replace(MBARI_DIR, "")
            context["thumbnail_fullrez_url"] = join(
                base_uri, "SeafloorMapping", img_path
            )
        except (AttributeError, ValueError):
            # no thumbnail_filename - return something
            context["thumbnail_fullrez_url"] = base_uri
        table = CompilationTable(
            Compilation.objects.filter(missions=mission),
            exclude=["missions"],
            order_by="-creation_date",
        )
        RequestConfig(self.request).configure(table)
        context["table"] = table
        return context

    def get_object(self):
        try:
            obj = super().get_object()
        except Mission.MultipleObjectsReturned:
            obj = (
                Mission.objects.prefetch_related("citations")
                .filter(slug=self.kwargs["slug"])
                .first()
            )
            messages.warning(
                self.request,
                f"Multiple Missions with slug '{self.kwargs['slug']}'. Using first one.",
            )
        return obj


class ExpeditionDetailView(DetailView):
    model = Expedition

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        expedition = super().get_object()

        table = MissionTable(
            Mission.objects.filter(expedition=expedition),
            exclude=["expedition"],
        )
        RequestConfig(self.request).configure(table)
        context["table"] = table

        return context

    def get_object(self):
        try:
            obj = super().get_object()
        except Expedition.MultipleObjectsReturned:
            obj = Expedition.objects.filter(slug=self.kwargs["slug"]).first()
            messages.warning(
                self.request,
                f"Multiple Expeditions with slug '{self.kwargs['slug']}'. Using first one.",
            )
        return obj


class CompilationDetailView(SuccessMessageMixin, DetailView):
    model = Compilation

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        compilation = self.get_object()
        try:
            context["thumbnail_url"] = compilation.thumbnail_image.url
        except (AttributeError, ValueError):
            context["thumbnail_url"] = join(
                settings.STATIC_URL, "images", "No_ZTopoSlopeNav_image.jpg"
            )
        try:
            site_uri = self.request.build_absolute_uri("/")
            base_uri = re.match(r"(http[s]*:\/\/[^:\/]*)", site_uri).group(1)
            img_path = compilation.thumbnail_filename.replace(MBARI_DIR, "")
            context["thumbnail_fullrez_url"] = join(
                base_uri, "SeafloorMapping", img_path
            )
        except (AttributeError, ValueError):
            # no thumbnail_filename - return something
            context["thumbnail_fullrez_url"] = base_uri

        table = MissionTable(
            Mission.objects.filter(compilations=compilation),
            exclude=["track_length", "area", "start_depth"],
            order_by="-start_date",
        )
        RequestConfig(self.request).configure(table)
        context["table"] = table
        return context

    def get_object(self):
        try:
            obj = super().get_object()
        except Compilation.MultipleObjectsReturned:
            obj = Compilation.objects.filter(slug=self.kwargs["slug"]).first()
            messages.warning(
                self.request,
                f"Multiple Compilations with slug '{self.kwargs['slug']}'. Using first one.",
            )
        return obj


class MissionSelectAPIView(View):
    """
    API endpoint to get missions filtered by spatial bounds and other filters.
    Returns JSON with mission data for display in results panel.
    """
    
    def get(self, request):
        try:
            # Get filter parameters
            filter_params = request.GET.copy()
            
            # Get spatial bounds
            xmin = filter_params.get('xmin')
            xmax = filter_params.get('xmax')
            ymin = filter_params.get('ymin')
            ymax = filter_params.get('ymax')
            
            if not all([xmin, xmax, ymin, ymax]):
                return JsonResponse({'error': 'Missing spatial bounds'}, status=400)
            
            # Validate and convert bounds to floats
            try:
                xmin_float = float(xmin)
                xmax_float = float(xmax)
                ymin_float = float(ymin)
                ymax_float = float(ymax)
            except (ValueError, TypeError):
                return JsonResponse({'error': 'Invalid spatial bounds coordinates'}, status=400)
            
            # Clamp to valid WGS84 range (handles floating-point precision from Leaflet
            # when drawing boxes with many missions). Reject only if wildly out of range.
            if (abs(xmin_float) > 360 or abs(xmax_float) > 360 or
                    not (-90 <= ymin_float <= 90) or not (-90 <= ymax_float <= 90)):
                return JsonResponse({'error': 'Coordinates out of valid range'}, status=400)
            xmin_float = max(-180, min(180, xmin_float))
            xmax_float = max(-180, min(180, xmax_float))
            ymin_float = max(-90, min(90, ymin_float))
            ymax_float = max(-90, min(90, ymax_float))
            
            # Validate bounding box ordering: xmin <= xmax and ymin <= ymax
            if xmin_float > xmax_float or ymin_float > ymax_float:
                return JsonResponse(
                    {'error': 'Invalid bounding box: xmin must be <= xmax and ymin must be <= ymax'},
                    status=400,
                )
            
            # Create polygon from bounds
            search_geom = Polygon(
                (
                    (xmin_float, ymin_float),
                    (xmin_float, ymax_float),
                    (xmax_float, ymax_float),
                    (xmax_float, ymin_float),
                    (xmin_float, ymin_float),
                ),
                srid=4326,
            )
            
            # Start with base queryset
            missions = Mission.objects.all()
            
            # Apply spatial filter: only missions whose nav_track or start_point is in the box.
            # Exclude grid_bounds — it can be larger than the track (full grid extent) and would
            # return missions whose tracks don't actually pass through the drawn box.
            missions = missions.filter(
                Q(nav_track__intersects=search_geom)
                | Q(start_point__within=search_geom)
            ).distinct()
            
            # Apply other filters if present (only when request has those filter keys)
            filter_type = filter_params.get('filter_type', '')
            mission_filter_keys = ['name', 'region_name', 'vehicle_name', 'platformtype', 'quality_categories', 'patch_test', 'repeat_survey', 'mgds_compilation', 'citation', 'citation_search', 'expedition__name']
            has_mission_filters = any(key in filter_params for key in mission_filter_keys)
            
            # Apply mission filter only when we have mission filter params (skip when only bbox/tmin/tmax)
            if has_mission_filters and (filter_type == 'mission' or filter_type == ''):
                # Pass request.GET so FilterSet form gets a proper QueryDict and validates
                mission_filter = MissionFilter(request.GET, queryset=missions)
                missions = mission_filter.qs
            
            # Apply expedition filter
            elif filter_type == 'expedition' and filter_params.get('name'):
                expedition_filter = ExpeditionFilter(filter_params, queryset=Expedition.objects.all())
                filtered_expeditions = expedition_filter.qs
                if filtered_expeditions.exists():
                    missions = missions.filter(expedition__in=filtered_expeditions)
            
            # Apply compilation filter
            elif filter_type == 'compilation' and filter_params.get('name'):
                compilation_filter = CompilationFilter(filter_params, queryset=Compilation.objects.all())
                filtered_compilations = compilation_filter.qs
                if filtered_compilations.exists():
                    missions = missions.filter(compilations__in=filtered_compilations).distinct()
            
            # Apply text search
            if filter_params.get('q'):
                search_string = filter_params.get('q')
                missions = missions.filter(
                    Q(name__icontains=search_string) 
                    | Q(notes_text__icontains=search_string)
                    | Q(expedition__name__icontains=search_string)
                )
            
            # Apply time filter
            tmin_str = filter_params.get('tmin')
            tmax_str = filter_params.get('tmax')
            if tmin_str and tmax_str:
                # Parse string dates to date objects to avoid TypeError with datetime.combine()
                min_date = parse_date(str(tmin_str))
                max_date = parse_date(str(tmax_str))
                if min_date and max_date:
                    missions = missions.filter(
                        start_date__gte=min_date,
                        end_date__lte=max_date,
                    )
            
            # Select related to avoid N+1 queries
            missions = missions.select_related('expedition').prefetch_related('quality_categories').order_by('start_date')
            
            # Serialize mission data
            mission_data = []
            for mission in missions:
                mission_data.append({
                    'slug': mission.slug,
                    'name': mission.name,
                    'start_date': mission.start_date.strftime('%Y-%m-%d') if mission.start_date else None,
                    'region_name': mission.region_name or None,
                    'track_length': str(mission.track_length) if mission.track_length else None,
                    'start_depth': str(mission.start_depth) if mission.start_depth else None,
                    'vehicle_name': mission.vehicle_name or None,
                    'expedition_name': mission.expedition.name if mission.expedition else None,
                })
            
            return JsonResponse({'missions': mission_data})
            
        except Exception as e:
            logging.error(f"Error in MissionSelectAPIView: {str(e)}")
            return JsonResponse({'error': str(e)}, status=500)


class MissionExportAPIView(View):
    """
    API endpoint to export missions as CSV or Excel.
    Accepts same filter parameters as MissionSelectAPIView.
    """
    
    def get(self, request):
        try:
            # Get filter parameters (same as MissionSelectAPIView)
            filter_params = request.GET.copy()
            export_format = filter_params.get('format', 'csv').lower()
            
            # Get spatial bounds
            xmin = filter_params.get('xmin')
            xmax = filter_params.get('xmax')
            ymin = filter_params.get('ymin')
            ymax = filter_params.get('ymax')
            
            if not all([xmin, xmax, ymin, ymax]):
                return HttpResponse('Missing spatial bounds', status=400)
            
            # Validate and convert bounds to floats
            try:
                xmin_float = float(xmin)
                xmax_float = float(xmax)
                ymin_float = float(ymin)
                ymax_float = float(ymax)
            except (ValueError, TypeError):
                return HttpResponse('Invalid spatial bounds coordinates', status=400)
            
            # Clamp to valid WGS84 range (handles floating-point precision from Leaflet).
            if (abs(xmin_float) > 360 or abs(xmax_float) > 360 or
                    not (-90 <= ymin_float <= 90) or not (-90 <= ymax_float <= 90)):
                return HttpResponse('Coordinates out of valid range', status=400)
            xmin_float = max(-180, min(180, xmin_float))
            xmax_float = max(-180, min(180, xmax_float))
            ymin_float = max(-90, min(90, ymin_float))
            ymax_float = max(-90, min(90, ymax_float))
            
            # Validate bounding box ordering: xmin <= xmax and ymin <= ymax
            if xmin_float > xmax_float or ymin_float > ymax_float:
                return HttpResponse(
                    'Invalid bounding box: xmin must be <= xmax and ymin must be <= ymax',
                    status=400,
                )
            
            # Create polygon from bounds
            search_geom = Polygon(
                (
                    (xmin_float, ymin_float),
                    (xmin_float, ymax_float),
                    (xmax_float, ymax_float),
                    (xmax_float, ymin_float),
                    (xmin_float, ymin_float),
                ),
                srid=4326,
            )
            
            # Start with base queryset
            missions = Mission.objects.all()
            
            # Apply spatial filter (same as MissionSelectAPIView: nav_track or start_point only)
            missions = missions.filter(
                Q(nav_track__intersects=search_geom)
                | Q(start_point__within=search_geom)
            ).distinct()
            
            # Apply other filters (same logic as MissionSelectAPIView)
            filter_type = filter_params.get('filter_type', '')
            mission_filter_keys = ['name', 'region_name', 'vehicle_name', 'platformtype', 'quality_categories', 'patch_test', 'repeat_survey', 'mgds_compilation', 'citation', 'citation_search', 'expedition__name']
            has_mission_filters = any(key in filter_params for key in mission_filter_keys)
            if has_mission_filters and (filter_type == 'mission' or filter_type == ''):
                mission_filter = MissionFilter(request.GET, queryset=missions)
                missions = mission_filter.qs
            elif filter_type == 'expedition' and filter_params.get('name'):
                expedition_filter = ExpeditionFilter(filter_params, queryset=Expedition.objects.all())
                filtered_expeditions = expedition_filter.qs
                if filtered_expeditions.exists():
                    missions = missions.filter(expedition__in=filtered_expeditions)
            elif filter_type == 'compilation' and filter_params.get('name'):
                compilation_filter = CompilationFilter(filter_params, queryset=Compilation.objects.all())
                filtered_compilations = compilation_filter.qs
                if filtered_compilations.exists():
                    missions = missions.filter(compilations__in=filtered_compilations).distinct()
            
            if filter_params.get('q'):
                search_string = filter_params.get('q')
                missions = missions.filter(
                    Q(name__icontains=search_string) 
                    | Q(notes_text__icontains=search_string)
                    | Q(expedition__name__icontains=search_string)
                )
            
            tmin_str = filter_params.get('tmin')
            tmax_str = filter_params.get('tmax')
            if tmin_str and tmax_str:
                # Parse string dates to date objects to avoid TypeError with datetime.combine()
                min_date = parse_date(str(tmin_str))
                max_date = parse_date(str(tmax_str))
                if min_date and max_date:
                    missions = missions.filter(
                        start_date__gte=min_date,
                        end_date__lte=max_date,
                    )
            
            missions = missions.select_related('expedition').prefetch_related('quality_categories').order_by('start_date')
            
            if export_format == 'csv':
                return self.export_csv(missions)
            elif export_format == 'excel':
                return self.export_excel(missions)
            else:
                return HttpResponse('Invalid format. Use csv or excel.', status=400)
                
        except Exception:
            logging.exception("Error in MissionExportAPIView")
            return HttpResponse('An internal error occurred. Please try again later.', status=500)
    
    def export_csv(self, missions):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="missions_export.csv"'
        
        writer = csv.writer(response)
        # Write header
        writer.writerow([
            'Name', 'Start Date', 'Region', 'Track Length (km)', 'Start Depth (m)',
            'Vehicle', 'Expedition', 'Quality Categories', 'Patch Test', 'Repeat Survey',
            'MGDS Compilation', 'Quality Comment'
        ])
        
        # Write data
        for mission in missions:
            quality_categories = ', '.join([qc.name for qc in mission.quality_categories.all()]) if mission.quality_categories.exists() else ''
            writer.writerow([
                mission.name or '',
                mission.start_date.strftime('%Y-%m-%d') if mission.start_date else '',
                mission.region_name or '',
                str(mission.track_length) if mission.track_length else '',
                str(mission.start_depth) if mission.start_depth else '',
                mission.vehicle_name or '',
                mission.expedition.name if mission.expedition else '',
                quality_categories,
                'Yes' if mission.patch_test else 'No',
                'Yes' if mission.repeat_survey else 'No',
                mission.mgds_compilation or '',
                mission.quality_comment or '',
            ])
        
        return response
    
    def export_excel(self, missions):
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
        except ImportError:
            # Fallback to CSV if openpyxl is not available
            return self.export_csv(missions)
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Missions"
        
        # Write header
        headers = [
            'Name', 'Start Date', 'Region', 'Track Length (km)', 'Start Depth (m)',
            'Vehicle', 'Expedition', 'Quality Categories', 'Patch Test', 'Repeat Survey',
            'MGDS Compilation', 'Quality Comment'
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Write data
        for row, mission in enumerate(missions, 2):
            quality_categories = ', '.join([qc.name for qc in mission.quality_categories.all()]) if mission.quality_categories.exists() else ''
            ws.cell(row=row, column=1, value=mission.name or '')
            ws.cell(row=row, column=2, value=mission.start_date.strftime('%Y-%m-%d') if mission.start_date else '')
            ws.cell(row=row, column=3, value=mission.region_name or '')
            ws.cell(row=row, column=4, value=mission.track_length if mission.track_length else '')
            ws.cell(row=row, column=5, value=mission.start_depth if mission.start_depth else '')
            ws.cell(row=row, column=6, value=mission.vehicle_name or '')
            ws.cell(row=row, column=7, value=mission.expedition.name if mission.expedition else '')
            ws.cell(row=row, column=8, value=quality_categories)
            ws.cell(row=row, column=9, value='Yes' if mission.patch_test else 'No')
            ws.cell(row=row, column=10, value='Yes' if mission.repeat_survey else 'No')
            ws.cell(row=row, column=11, value=mission.mgds_compilation or '')
            ws.cell(row=row, column=12, value=mission.quality_comment or '')
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = openpyxl.utils.get_column_letter(col)
            ws.column_dimensions[column_letter].width = 15
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="missions_export.xlsx"'
        
        return response
